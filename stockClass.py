# -*- coding: UTF-8 -*-

'''For the test purpose, unmark lines with '# for test' and mark '# formal'
new-stcok-fetching policy
1. run 'get_stockCodes.py'
2. accumulate new stocks and put them into stockCodes of stockClassInstance.py
3. run 'stockClassInstance.py retrieveStockData' with the parameter "" (python3 /home/pi/Python/pythonStockEvaluation/stockClassInstance.py retrieveStockData "")
4. run 'stockClassInstance.py computeStockKD' with the parameter "" (python3 /home/pi/Python/pythonStockEvaluation/stockClassInstance.py computeStockKD "")
'''

from fake_useragent import UserAgent
from lxml import etree
import collections
import copy
from abc import ABCMeta, abstractmethod
import requests
import pymysql.cursors
import json
import sys,os
import datetime
from fake_useragent import UserAgent
import time
import threading
from time import sleep,ctime
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Color
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import calendar
import matplotlib.ticker
import jieba
from operator import itemgetter
# from pylab import mpl
import glob
import matplotlib
import traceback

# print(matplotlib.matplotlib_fname())
#
# mpl.rcParams['font.sans-serif'] = ['SimHei'] #將預設字體改用SimHei字體

class stockClass(object):
    """This class is for retrieving stock-related data

    Attributes:
        timeout: webpage-fetching timeout
        threadAmount: thread number
    """

    __metaclass__ = ABCMeta

    # Static variables - Attributes of a class hold for all instances in all cases
	# globalTimeout = 5
	# globalThreadAmount = 5

    def __init__(self, timeout, threadAmount, training):
        self.timeout = timeout
        self.threadAmount = threadAmount
        self.training = training

    def getStockCodes(self, get_new=False):
        if get_new:
            sql = "SELECT stockCode FROM stocktable WHERE stockNewArrival='new'"
        else:
            sql = "SELECT stockCode FROM stocktable WHERE stockFinished='no'"
        try:
            # Execute the SQL command
            conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
            cursor = conn.cursor()
            cursor.execute(sql)
            # Fetch all the rows in a list of lists.
            stockCodes = []
            results = cursor.fetchall()
            for row in results:
                stockCodes.append(row[0])
            print(">>>>>Not Finished>>>>>")
            print(stockCodes)
            print("<<<<<Not Finished<<<<<")

            cursor.execute("UPDATE stocktable SET stockFinished='no'")

            cursor.execute(sql)
            # Fetch all the rows in a list of lists.
            stockCodes = []
            results = cursor.fetchall()
            for row in results:
                stockCodes.append(row[0])
        except:
            print("Error: unable to fecth data")

        cursor.close()
        conn.commit()
        conn.close()

        return stockCodes

    def getProxy(self, offset):
        # global globalTimeout

        training_proxy_sql = ''
        if self.training:
            training_proxy_sql = ' AND proxyUsedTimes=0'
            offset = 0
        else:
            import random
            random.seed(time.time())
            offset = random.randint(0, 500)

        while True:
            try:
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cur = conn.cursor()
                sql = """SELECT proxyIPPort FROM (
                    SELECT sid, proxyIPPort, proxyAvgReponseperiod, proxyUsedTimes, proxyFailtimes*proxyAvgReponseperiod AS formula FROM stockproxies) AS proxyFormula
                    WHERE proxyAvgReponseperiod<%s""" + training_proxy_sql + """ ORDER BY formula ASC, sid ASC LIMIT 1 OFFSET %s"""
                # sql = """SELECT proxyIPPort FROM (
                #     SELECT sid, proxyIPPort, proxyAvgReponseperiod, proxyFailtimes*proxyAvgReponseperiod AS formula FROM stockproxies) AS proxyFormula
                #     ORDER BY formula ASC, sid ASC LIMIT 1"""
                cur.execute(sql, (self.timeout, offset))
                proxy = cur.fetchone()

                if self.training and not proxy:
                    print('No unused proxies available. End training...')
                    os.kill(os.getpid(), 9)

                cur.close()
                conn.commit()
                conn.close()

                print('Using proxy=>' + proxy[0] + ' offset=>' + str(offset) + '\n\n')
                return proxy[0]
            except:
                print("Unexpected error:", sys.exc_info())
                cur.close()
                conn.close()

    def updateProxyInfo(self, proxy, succeedOrFail, executionTime):
        while True:
            try:
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cur = conn.cursor()
                sql = """UPDATE stockproxies
                    SET proxyAvgReponseperiod=(proxyAvgReponseperiod*proxyUsedTimes+%s)/(proxyUsedTimes+1),
    	                proxyUsedTimes=proxyUsedTimes+1,
    	                proxyFailtimes=proxyFailtimes+%s
                    WHERE proxyIPPort=%s"""
                cur.execute(sql, (executionTime if succeedOrFail else (executionTime + 10), 0 if succeedOrFail else 1, proxy))  # 10 is for penalty when failing

                cur.close()
                conn.commit()
                conn.close()

                return
            except:
                print("Unexpected error:", sys.exc_info())
                cur.close()
                conn.close()

    def update_stocktable_new_to_old(self):
        while True:
            try:
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cur = conn.cursor()
                sql = """UPDATE stocktable SET stockNewArrival=CONCAT('old_', CURRENT_TIMESTAMP) WHERE stockNewArrival='new'"""
                cur.execute(sql)  # 10 is for penalty when failing

                cur.close()
                conn.commit()
                conn.close()

                return
            except:
                print("Unexpected error:", sys.exc_info())
                cur.close()
                conn.close()

    def updateStockFinished(self, stock, status):
        while True:
            try:
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cur = conn.cursor()
                sql = """UPDATE stocktable SET stockFinished=%s WHERE stockcode=%s"""
                cur.execute(sql, (status, stock))

                cur.close()
                conn.commit()
                conn.close()

                return
            except:
                print("Unexpected error:", sys.exc_info())
                cur.close()
                conn.close()

    def getInstitutionalInvestors(self, date):
        # 自營商買賣超彙總表
        # http://www.twse.com.tw/fund/TWT43U?response=json&date=20180112
        # 投信買賣超彙總表
        # http://www.twse.com.tw/fund/TWT44U?response=json&date=&_=1515943335745
        # 外資及陸資買賣超彙總表
        # http://www.twse.com.tw/fund/TWT38U?response=json&date=20180112

        ua = UserAgent()
        insInvestArray = []

        hyphenDate = datetime.datetime.strptime(date, '%Y%m%d').strftime('%Y-%m-%d')

        # 自營商買賣超彙總表
        succeed = False
        nowProxy = self.getProxy(0)
        proxies = {"http": "http://" + nowProxy}
        while not succeed:
            try:
                header = {'User-Agent': str(ua.random)}
                res = requests.get("http://www.twse.com.tw/fund/TWT43U?response=json&date=" + str(date), headers=header, proxies=proxies, timeout=10)

                s = json.loads(res.text)
                if 'data' in s:
                    for data in s['data']:
                        # insInvestArray.append((hyphenDate, data[0].strip(), 'dealer', 'dealerSelf', data[2].strip().replace(",", "")))
                        # insInvestArray.append((hyphenDate, data[0].strip(), 'dealer', 'dealerSelf', ('' if data[3].strip().replace(",", "") == '0' else '-') + data[3].strip().replace(",", "")))
                        # insInvestArray.append((hyphenDate, data[0].strip(), 'dealer', 'dealerHedging', data[5].strip().replace(",", "")))
                        # insInvestArray.append((hyphenDate, data[0].strip(), 'dealer', 'dealerHedging', ('' if data[6].strip().replace(",", "") == '0' else '-') + data[6].strip().replace(",", "")))
                        insInvestArray.append((hyphenDate, data[0].strip(), 'dealer', 'dealer', data[8].strip().replace(",", "")))
                        insInvestArray.append((hyphenDate, data[0].strip(), 'dealer', 'dealer', ('' if data[9].strip().replace(",", "") == '0' else '-') + data[9].strip().replace(",", "")))
                    succeed = True
                elif 'stat' in s and '很抱歉' in s['stat']:
                    succeed = True
                else:
                    self.updateProxyInfo(nowProxy, False, 0)
                    nowProxy = self.getProxy(0)
                    proxies = {"http": "http://" + nowProxy}
            except:
                print("Unexpected error:", sys.exc_info())
                self.updateProxyInfo(nowProxy, False, 0)
                nowProxy = self.getProxy(0)
                proxies = {"http": "http://" + nowProxy}

        time.sleep(1)

        # 投信買賣超彙總表
        succeed = False
        nowProxy = self.getProxy(1)
        proxies = {"http": "http://" + nowProxy}
        while not succeed:
            try:
                header = {'User-Agent': str(ua.random)}
                res = requests.get("http://www.twse.com.tw/fund/TWT44U?response=json&date=" + str(date), headers=header, proxies=proxies, timeout=10)

                s = json.loads(res.text)
                if 'data' in s:
                    for data in s['data']:
                        insInvestArray.append((hyphenDate, data[1].strip(), 'investmentTrust', 'investmentTrust', data[3].strip().replace(",", "")))
                        insInvestArray.append((hyphenDate, data[1].strip(), 'investmentTrust', 'investmentTrust', ('' if data[4].strip().replace(",", "") == '0' else '-') + data[4].strip().replace(",", "")))
                    succeed = True
                elif 'stat' in s and '很抱歉' in s['stat']:
                    succeed = True
                else:
                    self.updateProxyInfo(nowProxy, False, 0)
                    nowProxy = self.getProxy(1)
                    proxies = {"http": "http://" + nowProxy}
            except:
                print("Unexpected error:", sys.exc_info())
                self.updateProxyInfo(nowProxy, False, 0)
                nowProxy = self.getProxy(1)
                proxies = {"http": "http://" + nowProxy}

        time.sleep(1)

        # 外資及陸資買賣超彙總表
        succeed = False
        nowProxy = self.getProxy(2)
        proxies = {"http": "http://" + nowProxy}
        while not succeed:
            try:
                header = {'User-Agent': str(ua.random)}
                res = requests.get("http://www.twse.com.tw/fund/TWT38U?response=json&date=" + str(date), headers=header, proxies=proxies, timeout=10)

                s = json.loads(res.text)
                if 'data' in s:
                    for data in s['data']:
                        insInvestArray.append((hyphenDate, data[1].strip(), 'foreignInvestor', 'foreignInvestor', data[9].strip().replace(",", "")))
                        insInvestArray.append((hyphenDate, data[1].strip(), 'foreignInvestor', 'foreignInvestor', ('' if data[10].strip().replace(",", "") == '0' else '-') + data[10].strip().replace(",", "")))
                    succeed = True
                elif 'stat' in s and '很抱歉' in s['stat']:
                    succeed = True
                else:
                    self.updateProxyInfo(nowProxy, False, 0)
                    nowProxy = self.getProxy(2)
                    proxies = {"http": "http://" + nowProxy}
            except:
                print("Unexpected error:", sys.exc_info())
                self.updateProxyInfo(nowProxy, False, 0)
                nowProxy = self.getProxy(2)
                proxies = {"http": "http://" + nowProxy}

        print(insInvestArray)
        if insInvestArray:  # not empty
            while True:
                try:
                    conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                    cur = conn.cursor()
                    insert = "INSERT IGNORE INTO stockInstitutionalInvestor (stockDate, stockCode, stockInvestorType, stockInvestorTypeDetail, stockAmount) VALUES (%s, %s, %s, %s, %s)"
                    cur.executemany(insert, insInvestArray)
                    cur.close()
                    conn.commit()
                    conn.close()
                    break
                except:
                    print("Unexpected error:", sys.exc_info())
                    cur.close()
                    conn.close()

    def transfer_Chinesedate_to_numericdate(self, input_date):
        input_date = re.sub('日.*', '', input_date)
        re_date = re.compile(r'(\d{2,3})年(\d{2})月(\d{2})$', re.S | re.UNICODE)
        for file_date in re_date.findall(input_date):
            file_date = str(int(file_date[0]) + 1911) + file_date[1] + file_date[2]
        # str(int(splitDate[0]) + 1911)
        return file_date

    def fetch_new_stocks(self):
        url = "http://pchome.megatime.com.tw/js/stock_list.js"
        r = requests.get(url)
        r.encoding = 'utf8'
        # print(r.text)

        pattern = re.compile(r'\(\s*\'(.*?)\'\s*,\s*\'(.*?)\'\s*,\s*\'(.*?)\'\s*,\s*\'(.*?)\'\s*\)')
        # match = pattern.match("['0050', '元大台灣50', '元大台灣50', ',0_00,']")

        stocksArray = []
        stocks = re.findall(r'\[\'(.*?)\'\s*,\s*\'(.*?)\'\s*,\s*\'(.*?)\'\s*,\s*\'(.*?)\'\s*\]', r.text)
        for stock in stocks:
            match = pattern.match(str(stock))
            stocksArray.append((match.group(1), match.group(2), match.group(3), match.group(4)))
            # print(match.group(1))
            # print(match.group(2))
            # print(match.group(3))
            # print(match.group(4))

        print(stocksArray)

        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        # conn.set_charset('utf8')

        cur = conn.cursor()
        # cur.execute("SELECT * FROM 膠片")
        insert = "INSERT IGNORE INTO stocktable (stockCode, stockName, stockFullname, stockNote) VALUES (%s, %s, %s, %s)"
        cur.executemany(insert, stocksArray)
        cur.close()
        conn.commit()
        conn.close()

    def fetch_new_stock_index_all(self):
        pass

    def retrieveStockDataFromFile(self, fetchDate):
        import urllib.request

        for f in glob.glob("stock_all_*"):
            os.remove(f)

        print('Begin downloading stock info...')
        url = 'http://www.twse.com.tw/exchangeReport/MI_INDEX?response=csv&date=' + fetchDate + '&type=ALL'
        while True:
            try:
                for f in glob.glob("stock_all_*"):
                    os.remove(f)
                urllib.request.urlretrieve(url, 'stock_all_' + fetchDate + '.csv')

                import platform
                if 'linux' in platform.platform().lower():
                    os.system("iconv -f big5 -t utf8 stock_all_" + fetchDate + ".csv -o stock_all_" + fetchDate + "_utf8.csv")
                break
            except:
                traceback.print_exc()
                sleep(1)

        stockCodes = self.getStockCodes()

        import csv
        start_fetch = False
        column_defined = False
        import platform
        if 'linux' in platform.platform().lower():
            csv_file_name = 'stock_all_' + fetchDate + '_utf8.csv'
        else:
            csv_file_name = 'stock_all_' + fetchDate + '.csv'

        with open(csv_file_name, newline='') as csvfile:
            rows = csv.reader(csvfile, delimiter=',')

            formalized_date = re.sub(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', fetchDate)

            # column_info = ['代號', '名稱', '股數', '收盤價']
            column_info = ['名稱', '股數', '收盤價']
            column_dict = {}
            stockDataArray = []
            for row in rows:
                str_row = str(row)
                # if start_fetch:
                #     print(str_row)
                if '收盤行情' in str_row and not start_fetch:
                    file_date = self.transfer_Chinesedate_to_numericdate(str_row)
                    if file_date!=fetchDate:
                        print('***ERROR*** File date does not correspond to fetching date.\n End fetching...\n')
                        os.kill(os.getpid(), 9)
                    start_fetch = True
                    continue
                elif start_fetch and not column_defined:
                    for index,column in enumerate(row):
                        for defined_column in column_info:
                            if defined_column in column:
                                column_defined = True
                                column_dict[index] = column
                elif start_fetch and column_defined:
                    row_stock_code = row[0].replace('="', '').replace('"', '')
                    for one_stock_code in stockCodes:
                        if one_stock_code == row_stock_code:
                            if not (row[8].isdigit() or row[8].replace('.', '', 1).isdigit()):
                                break
                            stockDataArray.append((one_stock_code, formalized_date, row[8], row[2]))

                    # for index, column in enumerate(row):
                    #     for one_stock_code in stockCodes:
                    #         if one_stock_code in column:

        if stockDataArray:  # not empty
            try:
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cur = conn.cursor()
                insert = "INSERT IGNORE INTO stockdata (stockCode, stockDate, stockIndex, stockVolume) VALUES (%s, %s, %s, %s)"
                cur.executemany(insert, stockDataArray)

                cur.close()
                conn.commit()
                conn.close()
            except:
                print("Unexpected error:", sys.exc_info())
                cur.close()
                conn.close()
        print(stockDataArray)

    def retrieveStockDataCounterFromFile(self, fetchDate, history=False):
        if history:
            from datetime import datetime, timedelta
            import urllib.request

            for f in glob.glob("stock_counter_all_*"):
                os.remove(f)

            stockCodes = self.getStockCodes()

            if not os.path.exists('counter_stocks'):
                os.makedirs('counter_stocks')

            nowProxy = self.getProxy(0)
            proxies = {"http": "http://" + nowProxy}

            while fetchDate!='20070101':
                print('Begin downloading stock info...')

                formalized_date = str(int(fetchDate[:4]) - 1911) + re.sub(r'(\d{4})(\d{2})(\d{2})', r'/\2/\3', fetchDate)
                url = 'http://www.tpex.org.tw/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php?l=zh-tw&o=csv&d=' + formalized_date + '&s=0,asc,0'
                print(url)
                while True:
                    try:
                        # create the object, assign it to a variable
                        proxy = urllib.request.ProxyHandler(proxies)
                        # construct a new opener using your proxy settings
                        opener = urllib.request.build_opener(proxy)
                        # install the openen on the module-level
                        urllib.request.install_opener(opener)
                        # make a request
                        urllib.request.urlretrieve(url, 'counter_stocks/stock_counter_all_' + fetchDate + '.csv')

                        import platform
                        if 'linux' in platform.platform().lower():
                            os.system("iconv -f big5 -t utf8 stock_counter_all_" + fetchDate + ".csv -o stock_counter_all_" + fetchDate + "_utf8.csv")
                        break
                    except:
                        traceback.print_exc()
                        self.updateProxyInfo(nowProxy, False, 0)
                        nowProxy = self.getProxy(0)
                        proxies = {"http": "http://" + nowProxy}

                '''
                import csv
                start_fetch = False
                column_defined = False
                import platform
                if 'linux' in platform.platform().lower():
                    csv_file_name = 'stock_counter_all_' + fetchDate + '_utf8.csv'
                else:
                    csv_file_name = 'stock_counter_all_' + fetchDate + '.csv'

                with open(csv_file_name, newline='') as csvfile:
                    rows = csv.reader(csvfile, delimiter=',')

                    formalized_date = re.sub(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', fetchDate)

                    # column_info = ['代號', '名稱', '股數', '收盤價']
                    column_info = ['名稱', '股數', '收盤']
                    column_dict = {}
                    stockDataArray = []
                    for row in rows:
                        str_row = str(row)
                        # if start_fetch:
                        #     print(str_row)
                        if '資料日期' in str_row and not start_fetch:
                            file_date = re.sub(r'[^\d\/]', r'', str_row)  # remove 資料日期
                            file_date = re.sub(r'(\d{3})/(\d{2})/(\d{2})', r'\1年\2月\3日', file_date)
                            file_date = self.transfer_Chinesedate_to_numericdate(file_date)

                            if file_date != fetchDate:
                                print('***ERROR*** File date does not correspond to fetching date.\n End fetching...\n')
                                os.kill(os.getpid(), 9)
                            start_fetch = True
                            continue
                        elif start_fetch and not column_defined:
                            for index, column in enumerate(row):
                                for defined_column in column_info:
                                    if defined_column in column:
                                        column_defined = True
                                        column_dict[index] = column
                        elif start_fetch and column_defined and row:
                            row_stock_code = row[0].replace('="', '').replace('"', '').strip()
                            for one_stock_code in stockCodes:
                                # print('===' + one_stock_code + '===' + row_stock_code + '===')
                                if one_stock_code == row_stock_code:
                                    # if row_stock_code=='6218':
                                    #     print(row_stock_code, row)
                                    if not (row[2].isdigit() or row[2].replace('.', '', 1).isdigit()):
                                        break
                                    stockDataArray.append((one_stock_code, formalized_date, row[2], row[8]))

                                    # for index, column in enumerate(row):
                                    #     for one_stock_code in stockCodes:
                                    #         if one_stock_code in column:

                if stockDataArray:  # not empty
                    try:
                        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                        cur = conn.cursor()
                        insert = "INSERT IGNORE INTO stockdata_copy (stockCode, stockDate, stockIndex, stockVolume) VALUES (%s, %s, %s, %s)"
                        cur.executemany(insert, stockDataArray)

                        cur.close()
                        conn.commit()
                        conn.close()
                    except:
                        print("Unexpected error:", sys.exc_info())
                        cur.close()
                        conn.close()
                print(stockDataArray)
                '''
                fetchDate = datetime.strptime(fetchDate, '%Y%m%d') - timedelta(days=1)
                fetchDate = fetchDate.strftime("%Y%m%d")
                sleep(10)

            exit(1)

        ################################
        import urllib.request

        for f in glob.glob("stock_counter_all_*"):
            os.remove(f)

        print('Begin downloading stock info...')
        formalized_date = str(int(fetchDate[:4]) - 1911) + re.sub(r'(\d{4})(\d{2})(\d{2})', r'/\2/\3', fetchDate)
        url = 'http://www.tpex.org.tw/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php?l=zh-tw&o=csv&d=' + formalized_date + '&s=0,asc,0'
        print(url)
        while True:
            try:
                for f in glob.glob("stock_counter_all_*"):
                    os.remove(f)
                urllib.request.urlretrieve(url, 'stock_counter_all_' + fetchDate + '.csv')

                import platform
                if 'linux' in platform.platform().lower():
                    os.system("iconv -f big5 -t utf8 stock_counter_all_" + fetchDate + ".csv -o stock_counter_all_" + fetchDate + "_utf8.csv")
                break
            except:
                traceback.print_exc()
                sleep(1)

        stockCodes = self.getStockCodes()

        import csv
        start_fetch = False
        column_defined = False
        import platform
        if 'linux' in platform.platform().lower():
            csv_file_name = 'stock_counter_all_' + fetchDate + '_utf8.csv'
        else:
            csv_file_name = 'stock_counter_all_' + fetchDate + '.csv'

        with open(csv_file_name, newline='') as csvfile:
            rows = csv.reader(csvfile, delimiter=',')

            formalized_date = re.sub(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', fetchDate)

            # column_info = ['代號', '名稱', '股數', '收盤價']
            column_info = ['名稱', '股數', '收盤']
            column_dict = {}
            stockDataArray = []
            for row in rows:
                str_row = str(row)
                # if start_fetch:
                #     print(str_row)
                if '資料日期' in str_row and not start_fetch:
                    file_date = re.sub(r'[^\d\/]', r'', str_row)  # remove 資料日期
                    file_date = re.sub(r'(\d{3})/(\d{2})/(\d{2})', r'\1年\2月\3日', file_date)
                    file_date = self.transfer_Chinesedate_to_numericdate(file_date)

                    if file_date!=fetchDate:
                        print('***ERROR*** File date does not correspond to fetching date.\n End fetching...\n')
                        os.kill(os.getpid(), 9)
                    start_fetch = True
                    continue
                elif start_fetch and not column_defined:
                    for index,column in enumerate(row):
                        for defined_column in column_info:
                            if defined_column in column:
                                column_defined = True
                                column_dict[index] = column
                elif start_fetch and column_defined and row:
                    row_stock_code = row[0].replace('="', '').replace('"', '').strip()
                    for one_stock_code in stockCodes:
                        # print('===' + one_stock_code + '===' + row_stock_code + '===')
                        if one_stock_code == row_stock_code:
                            # if row_stock_code=='6218':
                            #     print(row_stock_code, row)
                            if not (row[2].isdigit() or row[2].replace('.', '', 1).isdigit()):
                                break
                            stockDataArray.append((one_stock_code, formalized_date, row[2], row[8]))

                    # for index, column in enumerate(row):
                    #     for one_stock_code in stockCodes:
                    #         if one_stock_code in column:

        if stockDataArray:  # not empty
            try:
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cur = conn.cursor()
                insert = "INSERT IGNORE INTO stockdata (stockCode, stockDate, stockIndex, stockVolume) VALUES (%s, %s, %s, %s)"
                cur.executemany(insert, stockDataArray)

                cur.close()
                conn.commit()
                conn.close()
            except:
                print("Unexpected error:", sys.exc_info())
                cur.close()
                conn.close()
        print(stockDataArray)

    def insertStockDataCounterFromFile(self, fetchDate, folder_name='counter_stocks'):
        from os import listdir
        from os.path import isfile, join
        onlyfiles = [f for f in listdir(folder_name) if isfile(join(folder_name, f))]

        stockCodes = self.getStockCodes()

        for csv_file_name in onlyfiles:
            import csv
            start_fetch = False
            column_defined = False
            fetchDate = csv_file_name.replace('stock_counter_all_', '').replace('.csv', '')

            # import platform
            # if 'linux' in platform.platform().lower():
            #     csv_file_name = 'stock_counter_all_' + fetchDate + '_utf8.csv'
            # else:
            #     csv_file_name = 'stock_counter_all_' + fetchDate + '.csv'

            with open(folder_name + '/' + csv_file_name, newline='') as csvfile:
                rows = csv.reader(csvfile, delimiter=',')

                formalized_date = re.sub(r'(\d{4})(\d{2})(\d{2})', r'\1-\2-\3', fetchDate)

                # column_info = ['代號', '名稱', '股數', '收盤價']
                column_info = ['名稱', '股數', '收盤']
                column_dict = {}
                stockDataArray = []
                for row in rows:
                    str_row = str(row)
                    # if start_fetch:
                        # print(str_row)
                    if '資料日期' in str_row and not start_fetch:
                        file_date = re.sub(r'[^\d\/]', r'', str_row)  # remove 資料日期
                        file_date = re.sub(r'(\d{2,3})/(\d{2})/(\d{2})', r'\1年\2月\3日', file_date)
                        file_date = self.transfer_Chinesedate_to_numericdate(file_date)

                        if file_date != fetchDate:
                            print('***ERROR*** File date does not correspond to fetching date.\n End fetching...\n')
                            os.kill(os.getpid(), 9)
                        start_fetch = True
                        continue
                    elif start_fetch and not column_defined:
                        for index, column in enumerate(row):
                            for defined_column in column_info:
                                if defined_column in column:
                                    column_defined = True
                                    column_dict[index] = column
                    elif start_fetch and column_defined and row:
                        row_stock_code = row[0].replace('="', '').replace('"', '').strip()
                        for one_stock_code in stockCodes:
                            # print('===' + one_stock_code + '===' + row_stock_code + '===')
                            if one_stock_code == row_stock_code:
                                # if row_stock_code=='6218':
                                #     print(row_stock_code, row)
                                if not (row[2].isdigit() or row[2].replace('.', '', 1).isdigit()):
                                    break
                                stockDataArray.append((one_stock_code, formalized_date, row[2], row[8]))

                                # for index, column in enumerate(row):
                                #     for one_stock_code in stockCodes:
                                #         if one_stock_code in column:

            if stockDataArray:  # not empty
                try:
                    conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                    cur = conn.cursor()
                    insert = "INSERT IGNORE INTO stockdata (stockCode, stockDate, stockIndex, stockVolume) VALUES (%s, %s, %s, %s)"
                    cur.executemany(insert, stockDataArray)

                    cur.close()
                    conn.commit()
                    conn.close()
                except:
                    print("Unexpected error:", sys.exc_info())
                    cur.close()
                    conn.close()
            print(stockDataArray)


    #SELECT stockDate, COUNT(stockDate) FROM stockdata GROUP BY stockDate ORDER BY stockDate DESC
    def retrieveStockData(self, stockCodes, fromCode, endCode, offset, fetchDate):
        # global globalTimeout
        updateVolume = False

        print('>>>>>', fromCode, ctime(), '<<<<<')

        nowProxy = self.getProxy(offset)
        proxies = {"http": "http://" + nowProxy}
        now = datetime.datetime.now()

        if fetchDate != "":
            fromYear = int(fetchDate[0:4])
            endYear = int(fetchDate[0:4]) - 1
            fromMonth = int(fetchDate[4:6])
            endMonth = int(fetchDate[4:6]) - 1
        else:
            fromYear = now.year
            endYear = 1998
            fromMonth = now.month
            endMonth = 0

        i = fromCode  # this is the list index of one stock
        add_num = 0 if self.training else 1
        while i < endCode:
            # for stock in stockCodes:
            stock = str(stockCodes[i])
            self.updateStockFinished(stock, 'ing')
            finish = False
            # for year in range(now.year, 1998, -1):
            for year in range(fromYear, endYear, -1):
                if finish:
                    break

                if fetchDate == "":
                    fromMonth = 12
                    if year == now.year:
                        fromMonth = now.month

                for month in range(fromMonth, endMonth, -1):
                    if finish:
                        break

                    date = str(year) + str(month).zfill(2) + '01'

                    # url_twse = 'http://www.twse.com.tw/exchangeReport/STOCK_DAY_AVG?response=json&date=' + date + '&stockNo=' + stock
                    url_twse = 'http://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date=' + date + '&stockNo=' + stock
                    print(url_twse)

                    succeed = False
                    while not succeed:
                        ua = UserAgent()
                        header = {'User-Agent': str(ua.random)}

                        # fetchSucceed = False
                        # while not fetchSucceed:
                        while True:
                            try:
                                # if self.training:
                                if True:
                                    nowProxy = self.getProxy(offset)
                                    proxies = {"http": "http://" + nowProxy}

                                start = time.time()
                                res = requests.get(url_twse, headers=header, proxies=proxies, timeout=self.timeout)
                                end = time.time()

                                # if (end - start) >= self.timeout:
                                #     nowProxy = self.getProxy(offset)
                                #     proxies = {"http": "http://" + nowProxy}

                                print(stock + '===>' + res.text)
                                s = json.loads(res.text)
                                # fetchSucceed = True

                                if 'data' in s:
                                    self.updateProxyInfo(nowProxy, True, end-start)
                                    succeed = True # this is for this round
                                    break
                                elif 'stat' in s and '很抱歉' in s['stat']:
                                    self.updateProxyInfo(nowProxy, True, end-start)
                                    succeed = True  # this is for this round
                                    finish = True  # this is for this stock
                                    break
                                else:
                                    self.updateProxyInfo(nowProxy, False, 0)
                                    nowProxy = self.getProxy(offset)
                                    proxies = {"http": "http://" + nowProxy}
                            except:
                                print("Unexpected error:", sys.exc_info())
                                self.updateProxyInfo(nowProxy, False, 0)
                                nowProxy = self.getProxy(offset)
                                proxies = {"http": "http://" + nowProxy}

                        stockDataArray = []
                        if succeed and 'data' in s:
                            for data in (s['data']):
                                if '月' not in data[0]:
                                    splitDate = data[0].split('/')
                                    splitDate[0] = str(int(splitDate[0]) + 1911)
                                    splitDate = '-'.join(splitDate)

                                    # if not (data[1].isdigit() or data[1].replace('.', '', 1).isdigit()):
                                    if not (data[6].isdigit() or data[6].replace('.', '', 1).isdigit()):
                                        continue

                                    if fetchDate != "" and splitDate.replace('-', '') == fetchDate:
                                        # stockDataArray.append((stock, splitDate, data[1]))
                                        stockDataArray.append((stock, splitDate, data[6], data[1]))
                                    elif fetchDate == "":
                                        # stockDataArray.append((stock, splitDate, data[1]))
                                        if not updateVolume:
                                            stockDataArray.append((stock, splitDate, data[6], data[1]))
                                        else:
                                            stockDataArray.append((data[1], stock, splitDate))

                            print(stockDataArray)
                            # succeed = True

                            if stockDataArray:  # not empty
                                try:
                                    conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                                    cur = conn.cursor()
                                    if not updateVolume:
                                        insert = "INSERT IGNORE INTO stockdata (stockCode, stockDate, stockIndex, stockVolume) VALUES (%s, %s, %s, %s)"
                                        cur.executemany(insert, stockDataArray)
                                    else:
                                        cur.executemany("UPDATE stockdata SET stockVolume=%s WHERE stockCode=%s AND stockDate=%s", stockDataArray)
                                    cur.close()
                                    conn.commit()
                                    conn.close()
                                except:
                                    print("Unexpected error:", sys.exc_info())
                                    cur.close()
                                    conn.close()

                        # if 'stat' in s and '很抱歉' in s['stat']:
                        #     succeed = True
                        #     finish = True
                            # updateStockFinished(stock, 'yes')
                        if not succeed:
                            print("Fail: " + url_twse + " . Trying...")

                        time.sleep(1)

            self.updateStockFinished(stock, 'yes')
            i += add_num

    def computeStockKD(self, stockCodes, fromCode, endCode, offset, fetchDate):
        print('#####', fromCode, ctime(), '#####')

        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cursor = conn.cursor()

        for i in range(fromCode, endCode):
            # for stock in stockCodes:
            stock = stockCodes[i]
            stockData = []

            sqlSucceed = False
            while not sqlSucceed:
                if fetchDate == "":
                    sql = "SELECT * FROM stockdata WHERE stockcode=%s ORDER BY stockDate ASC"
                else:
                    sql = "SELECT * FROM stockdata WHERE stockcode=%s AND stockdate<=%s ORDER BY stockDate DESC LIMIT 9"

                try:
                    if fetchDate=='':
                        cursor.execute(sql, (stock))
                    else:
                        hyphenDate = datetime.datetime.strptime(fetchDate, '%Y%m%d').strftime('%Y-%m-%d')
                        cursor.execute(sql, (stock, hyphenDate))
                    # Fetch all the rows in a list of lists.
                    results = cursor.fetchall()
                    for row in results:
                        stockData.append([str(row[2]), row[3], row[4], row[5]]) #stockDate, stockIndex, stockK,stockD
                    sqlSucceed = True
                except:
                    print("Error: unable to fecth data")

            if len(stockData) < 9:
                continue

            if fetchDate != "":
                stockData = stockData[::-1] # reverse the list since we only need the current date
                end = 9
            else:
                end = len(stockData)

            print(stockData)

            curK = curD = -1.0
            stockKDList = []
            for j in range(8, end):
                if stockData[j][2]!=0.0 and stockData[j][3]!=0.0:   # if computed before, skip it
                    break

                if curK!=-1.0 and curD!=-1.0:
                    preK = curK
                    preD = curD
                elif stockData[j - 1][2] == 0.0 and stockData[j - 1][3] == 0.0:
                    preK = 50.0
                    preD = 50.0
                else:
                    preK = stockData[j - 1][2]
                    preD = stockData[j - 1][3]

                slice = stockData[j - 8:j + 1]
                print(stock, slice)
                print(stockData[j])
                print(max(slice, key=lambda x: x[1]))
                print(min(slice, key=lambda x: x[1]))
                if (float((max(slice, key=lambda x: x[1]))[1]) - float(min(slice, key=lambda x: x[1])[1])) != 0:
                    rsv = (float(stockData[j][1]) - float(min(slice, key=lambda x: x[1])[1])) / \
                        (float((max(slice, key=lambda x: x[1]))[1]) - float(min(slice, key=lambda x: x[1])[1])) * 100.0
                    curK = 2.0 / 3.0 * preK + 1.0 / 3.0 * rsv
                    curD = 2.0 / 3.0 * preD + 1.0 / 3.0 * curK
                else:
                    # rsv = 50.0
                    curK = preK
                    curD = preD
                print(rsv, preK, preD)
                print((curK, curD, stock, stockData[j][0]))
                stockKDList.append((curK, curD, stock, stockData[j][0]))

            while True:
                try:
                    sql = "UPDATE stockdata SET stockK=%s, stockD=%s WHERE stockCode=%s AND stockDate=%s"
                    cursor.executemany(sql, stockKDList)
                    conn.commit()
                    break
                except:
                    print("Unexpected error:", sys.exc_info())

        cursor.close()
        conn.close()

    def computeStockMA(self, stockCodes, fromCode, endCode, offset, fetchDate):
        print('#####', fromCode, ctime(), '#####')

        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cursor = conn.cursor()

        for stockI in range(fromCode, endCode):
            # for stock in stockCodes:
            stock = stockCodes[stockI]
            stockData = []

            sqlSucceed = False
            hyphenDate = datetime.datetime.strptime(fetchDate, '%Y%m%d').strftime('%Y-%m-%d')
            while not sqlSucceed:
                sql = "SELECT * FROM stockdata WHERE stockcode=%s AND stockdate<=%s ORDER BY stockDate DESC LIMIT 50"

                try:
                    cursor.execute(sql, (stock,hyphenDate))
                    # Fetch all the rows in a list of lists.
                    results = cursor.fetchall()
                    for row in results:
                        stockData.append([str(row[2]), row[3]]) #stockDate, stockIndex
                    sqlSucceed = True
                except:
                    print("Error: unable to fecth data")

            # if len(stockData) < 9:
            #     continue

            #if fetchDate != "":
            #    stockData = stockData[::-1] # reverse the list since we only need the current date
            #     end = 9
            # else:
            #     end = len(stockData)

            print(stockData)

            maList = [18, 50]
            maIndices = []
            for ma in maList:
                if len(stockData) >= ma:
                    sum = 0.0
                    for maI in range(0,ma):
                        sum += float(stockData[maI][1])

                    maIndices.append(sum/float(ma))
                else:
                    maIndices.append(0.0)

            while True:
                try:
                    sql = "UPDATE stockdata SET "
                    for ma in maList:
                        sql += "stockMA" + str(ma) + "=%s,"
                    sql = sql[:-1]  # remove the last comma
                    sql += " WHERE stockCode=%s AND stockDate=%s"
                    maIndices.append(stock)
                    maIndices.append(hyphenDate)

                    cursor.execute(sql, (maIndices))
                    conn.commit()
                    break
                except:
                    print("Unexpected error:", sys.exc_info())

        cursor.close()
        conn.close()

    def getStockNameInfoStartupdate(self):
        sql = """
        SELECT st.stockCode, st.stockName, st.stockInfo, stStartup.stockdate FROM stocktable AS st
        LEFT JOIN (
        SELECT stockcode, stockdate FROM stockdata GROUP BY stockcode ORDER BY stockdate ASC
        ) stStartup ON st.stockCode=stStartup.stockcode
        """
        try:
            # Execute the SQL command
            conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
            cursor = conn.cursor()
            cursor.execute(sql)
            # Fetch all the rows in a list of lists.
            stockCodeNames = {}
            results = cursor.fetchall()
            for row in results:
                stockCodeNames[row[0]] = [row[1],row[2], str(row[3])]
        except:
            print("Error: unable to fecth data")

        cursor.close()
        conn.commit()
        conn.close()

        return stockCodeNames

    def getStockInfo(self):
        stockCodes = self.getStockCodes()
        # stockCodes = ['9962']

        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cur = conn.cursor()
        sql = "UPDATE stocktable SET stockInfo=%s WHERE stockcode=%s"

        for stock in stockCodes:
            url_yahoo = 'https://tw.finance.yahoo.com/d/s/company_' + stock + '.html'
            print(url_yahoo)

            while True:
                nowProxy = self.getProxy(0)
                proxies = {"http": "http://" + nowProxy}
                ua = UserAgent()
                header = {'User-Agent': str(ua.random)}

                try:
                    res = requests.get(url_yahoo, headers=header, proxies=proxies, timeout=self.timeout)
                    res.encoding = 'big5-hkscs'
                    if '奇摩股市' in res.text:
                        print('>>>Found<<<')
                        print(res.text)

                        reGetStockInfo = re.compile(r'td\ width.*?營收比重.*?yui-td-left">(.*?)<\/td>', re.S|re.UNICODE)
                        for stockInfo in reGetStockInfo.findall(res.text):
                            print(stock,stockInfo)
                            while True:
                                try:
                                    cur.execute(sql, (stockInfo.strip(),stock))
                                    conn.commit()
                                    break
                                except:
                                    print("Unexpected error:", sys.exc_info())
                        break
                    elif 'tw.yahoo.com/?err=404' or '奇摩首頁' or '內容目前不可用' in res.text:
                        print('>>>Not Found<<<')
                        print(res.text)
                        break
                except:
                    print("Unexpected error:", sys.exc_info())

        cur.close()
        conn.close()

    def move_today_potential_to_yesterday_potential(self):
        # Move stockIsTodayPotential to stockIsYesterdayPotential of stocktable
        while True:
            try:
                # Execute the SQL command
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cursor = conn.cursor()
                cursor.execute('UPDATE stocktable SET stockIsYesterdayPotential=stockIsTodayPotential')
                cursor.execute('UPDATE stocktable SET stockIsTodayPotential=0')
                break
            except:
                print("Error: unable to fecth data")

        cursor.close()
        conn.commit()
        conn.close()

    def retrieveLowestIndexCurrentIndex(self,date):
        #*** Get lowest indices
        sql = """SELECT sd.stockcode, sd.stockdate, sd.stockindex, sd.stockK, sd.stockD, sd.stockMA18, sd.stockMA50
                FROM (
                SELECT stockcode, MIN(NULLIF(stockindex, 0)) AS minindex
                FROM stockdata
                GROUP BY stockcode
                ) AS mindata INNER JOIN stockdata AS sd ON sd.stockcode=mindata.stockcode AND sd.stockindex=mindata.minindex"""

        while True:
            try:
                # Execute the SQL command
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cursor = conn.cursor()
                cursor.execute(sql)
                # Fetch all the rows in a list of lists.
                stockCodeDateLowestindex = {}
                results = cursor.fetchall()
                for row in results:
                    stockCodeDateLowestindex[row[0]] = [str(row[1]), row[2], row[5], row[6], row[3], row[4]]

                print(stockCodeDateLowestindex)
                break
            except:
                print("Error: unable to fecth data")
        #___ Get lowest indices

        self.move_today_potential_to_yesterday_potential()

        #*** Get current indices
        sql = """SELECT * FROM stockdata sd INNER JOIN (
	              SELECT stockCode, MAX(stockDate) AS maxDate FROM stockdata GROUP BY stockCode
                 ) stockMaxDate ON sd.stockCode=stockMaxDate.stockCode AND sd.stockDate=stockMaxDate.maxDate"""

        while True:
            try:
                # Execute the SQL command
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cursor = conn.cursor()
                cursor.execute(sql)
                # Fetch all the rows in a list of lists.
                stockCodeCurrentindex = {}
                results = cursor.fetchall()
                for row in results:
                    stockCodeCurrentindex[row[1]] = [str(row[2]), row[3], row[4], row[5]]

                print(stockCodeCurrentindex)
                break
            except:
                print("Error: unable to fecth data")
        #___ Get current indices

        #*********** Find potential stocks
        #   1. Select the lowest price of one stock and compare to the current index
        #   2. if the ratio of the absolute value of the difference between the lowest and the current to the lowest is within 10%
        #   3. also pick the KD values these ten days
        #***********
        # sqlGetLastDays14 = 'SELECT stockcode,stockdate,stockindex,stockK,stockD FROM stockdata WHERE stockcode=%s ORDER BY stockdate DESC LIMIT 7'
        # sqlGetLastSevenDaysInvest = 'SELECT stockcode,stockdate,SUM(stockAmount) FROM stockinstitutionalinvestor WHERE stockcode=%s GROUP BY stockdate ORDER BY stockdate DESC LIMIT 7'
        sqlGetLastDays14 = """
            SELECT stockindices.stockcode,stockindices.stockdate,stockindices.stockindex,stockindices.stockK,stockindices.stockD,sumSA.sSA,stockindices.stockMA18,stockindices.stockMA50,stockindices.stockVolume FROM
                (SELECT stockcode,stockdate,stockindex,stockK,stockD,stockMA18,stockMA50,stockVolume FROM stockdata WHERE stockcode=%s 
                ORDER BY stockdate DESC LIMIT 14
                ) stockindices LEFT JOIN (
                SELECT stockcode,stockdate,SUM(stockAmount) AS sSA FROM stockinstitutionalinvestor WHERE stockcode=%s 
                GROUP BY stockdate ORDER BY stockdate DESC LIMIT 14) AS sumSA 
            ON stockindices.stockcode=sumSA.stockcode AND stockindices.stockdate=sumSA.stockdate
        """
        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cursor = conn.cursor()
        stockCodeIndices = {}
        # whiteList = ['0050','2634','2722','3057','3356','4141','5519','5706','8072','8429']
        # whiteList = ['2324','8429']
        whiteList = ['0050', '2330', '2412', '2105', '2596', '2823', '2885', '2891', '6218']
        for stock in stockCodeCurrentindex:
            try:
                if abs(float(stockCodeCurrentindex[stock][1])-float(stockCodeDateLowestindex[stock][1]))/float(stockCodeDateLowestindex[stock][1]) < 0.2 or stock in whiteList:  # formal
                # if stock in whiteList:  # for test
                    while True:
                        try:
                            # Execute the SQL command
                            cursor.execute(sqlGetLastDays14, (stock,stock))
                            # Fetch all the rows in a list of lists.
                            results = cursor.fetchall()
                            if results:
                                stockCodeIndices[stock] = []
                                for row in results:
                                    volume = int(re.sub(r"[^\d]", "", row[8])) if len(re.sub(r"[^\d]", "", row[8])) > 0 else 0
                                    stockCodeIndices[stock].append([str(row[1]), row[2], row[6], row[7], row[3], row[4], row[5], volume])
                                    #stockcode=>stockdate,stockindex,stockMA18,stockMA50,stockK,stockD,amount
                                stockCodeIndices[stock].append([str(stockCodeDateLowestindex[stock][0]+'(LOWEST)'), stockCodeDateLowestindex[stock][1], stockCodeDateLowestindex[stock][2], stockCodeDateLowestindex[stock][3], stockCodeDateLowestindex[stock][4], stockCodeDateLowestindex[stock][5]])
                            break
                        except:
                            print("Unexpected error:", sys.exc_info())

                    # print(">>>")
                    # print(stock, stockCodeCurrentindex[stock])
                    # print(stock, stockCodeDateLowestindex[stock])
                    # print("<<<")
            except:
                print("Unexpected error:", sys.exc_info())
        print(stockCodeIndices)
        #___ Find potential stocks

        # copy stockCodeIndices since stockCodeIndices is auto-appended (weird)
        stockCodeIndicesBackup = copy.deepcopy(stockCodeIndices)

        #*** Output to a CSV file
        for f in glob.glob("potentialStocks-*"):
            os.remove(f)
        now = datetime.datetime.now()
        file = open('potentialStocks-'+now.strftime("%Y-%m-%d")+'.csv', 'w', newline='')
        csvCursor = csv.writer(file)

        # write header to csv file
        csvHeader = ['stockCode', 'stockName', 'stockInfo', 'stockDate', 'stockIndex', 'stockMA18','stockMA50', 'stockK', 'stockD', 'amount']
        csvCursor.writerow(csvHeader)
        line = []
        stockCodeNames = self.getStockNameInfoStartupdate()

        for i in range(len(csvHeader)):
            line.append('----------')
        csvCursor.writerow(line)
        for stock in stockCodeIndices:
            first = True
            print(stock)

            d1 = datetime.date.today()
            d0 = datetime.date(int(stockCodeNames[stock][2].split("-")[0]), int(stockCodeNames[stock][2].split('-')[1]), int(stockCodeNames[stock][2].split('-')[2]))
            delta = (d1 - d0).days
            for oneRow in stockCodeIndices[stock]:
                #stockcode=>stockdate,stockindex,stockK,stockD,amount
                if first and (oneRow[2]>30 or oneRow[3]>30 or re.search('[a-zA-Z]', stock) or delta < 365*5):
                    if stock not in whiteList:
                        break

                oneRow.insert(0, stock)
                oneRow.insert(1, stockCodeNames[stock][0])
                oneRow.insert(2, stockCodeNames[stock][1])
                if first:
                    oneRow.append('https://tw.stock.yahoo.com/q/ta?s=' + stock + '&tech_submit=%ACd+%B8%DF')
                    first = False
                csvCursor.writerow(oneRow)

            if not first:
                csvCursor.writerow([stock, stockCodeNames[stock][0],stockCodeNames[stock][1],str(stockCodeNames[stock][2]) + ' 上市'])
                csvCursor.writerow([])

        file.close()
        #___ Output to a CSV file

        stockCodeIndices = copy.deepcopy(stockCodeIndicesBackup)

        #*** Output to a Excel file
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        ws.title = 'potentialStocks'  # 設置worksheet的標題
        ft = Font()
        ft.underline = 'single'    # add single underline
        ft.color = Color(rgb='000000FF')  # add blue color
        ft2 = Font()
        ft2.color = Color(rgb='000000FF')  # add blue color

        # write header
        excelHeader = ['stockCode', 'stockName', 'stockInfo', 'stockDate', 'stockIndex','stockMA18','stockMA50', 'stockK', 'stockD', 'amount']
        # ws.append(excelHeader)
        line = []
        for i in range(len(excelHeader)):
            line.append('----------')
        # ws.append(line)

        qualifiedStocks = []
        stockCodeNames = self.getStockNameInfoStartupdate()
        for stock in stockCodeIndices:
            first = True
            print(stock)

            d1 = datetime.date.today()
            d0 = datetime.date(int(stockCodeNames[stock][2].split("-")[0]), int(stockCodeNames[stock][2].split('-')[1]), int(stockCodeNames[stock][2].split('-')[2]))
            delta = (d1 - d0).days
            for oneRow in stockCodeIndices[stock]:
                if first and (oneRow[2]>30 or oneRow[3]>30 or re.search('[a-zA-Z]', stock) or delta < 365*5):
                    if stock not in whiteList:
                        break
                    else:
                        ws.append(excelHeader)
                        ws.append(line)
                elif first:
                    ws.append(excelHeader)
                    ws.append(line)

                #For drawing usage
                qualifiedStocks.append(stock)

                oneRow.insert(0, stock)
                oneRow.insert(1, stockCodeNames[stock][0])
                oneRow.insert(2, stockCodeNames[stock][1])
                ws.append(oneRow)
                if stock in whiteList:
                    ws['B'+str(ws._current_row)].font = ft2
                if first:
                    ws['K'+str(ws._current_row)] = '=HYPERLINK("{}", "{}")'.format('https://tw.stock.yahoo.com/q/ta?s=' + stock + '&tech_submit=%ACd+%B8%DF', "Link")
                    ws['K'+str(ws._current_row)].font = ft
                    first = False

            if not first:
                ws['D'+str(ws._current_row)].font = ft2
                ws['E'+str(ws._current_row)].font = ft2
                ws['F'+str(ws._current_row)].font = ft2
                ws['G'+str(ws._current_row)].font = ft2
                ws['H'+str(ws._current_row)].font = ft2
                ws['I'+str(ws._current_row)].font = ft2
                ws.append([stock, stockCodeNames[stock][0],stockCodeNames[stock][1],str(stockCodeNames[stock][2]) + ' 上市'])
                ws['D'+str(ws._current_row)].font = ft2
                if stock in whiteList:
                    ws['B'+str(ws._current_row)].font = ft2
                ws.append([])

        wb.save('potentialStocks-'+datetime.datetime.now().strftime("%Y-%m-%d")+'.xlsx')
        #___ Output to a Excel file

        stockCodeIndices = copy.deepcopy(stockCodeIndicesBackup)

        #*** Draw diagrams
        for f in glob.glob("flask-stock-decken/static/stockDrawing_*.png"):
            os.remove(f)

        days = 14
        x = []
        for i in range(1,days+1):
            x.append(i)
        x = np.asarray(x) # since zip doesn't accept list

        first = True
        # qualifiedStocks = ['0050','5706','9928','911868']
        for stock in qualifiedStocks:
            plt.rcParams["figure.figsize"] = [12, 6]    # set figure size to enlarge the plot. Remember to do >>> \includegraphics[width=1.0\textwidth] <<<
            fig, (subplot1, subplot2) = plt.subplots(2, 1,gridspec_kw = {'height_ratios':[5, 1]})#)
            subplot1.set_title(stock + '_' + stockCodeNames[stock][0] + '(' + stockCodeNames[stock][1][:70] + ')')

            stockInfoDict = {}
            stockInfoDict['date'] = []
            stockInfoDict['index'] = []
            stockInfoDict['ma18'] = []
            stockInfoDict['ma50'] = []
            stockInfoDict['k'] = []
            stockInfoDict['d'] = []
            stockInfoDict['amount'] = []
            stockInfoDict['dailyAmount'] = []
            for oneRow in stockCodeIndices[stock]:
                # print(oneRow)
                stockInfoDict['date'].append(oneRow[0])
                # print('111')
                stockInfoDict['index'].append(oneRow[1])
                # print('222')
                stockInfoDict['ma18'].append(oneRow[2])
                # print('333')
                stockInfoDict['ma50'].append(oneRow[3])
                # print('444')
                stockInfoDict['k'].append(oneRow[4])
                # print('555')
                stockInfoDict['d'].append(oneRow[5])
                # print('666')
                try:
                    stockInfoDict['amount'].append(oneRow[6]) if oneRow[6] else stockInfoDict['amount'].append(0)
                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)

                try:
                    stockInfoDict['dailyAmount'].append(oneRow[7])
                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)

            # print(stockInfoDict)

            xtickLabels = stockInfoDict['date'][::-1][1:] # remove the lowest info
            lowestIndex = stockInfoDict['index'][-1]
            stockInfoDict['index'] = stockInfoDict['index'][::-1][1:] # remove the lowest info
            stockInfoDict['ma18'] = stockInfoDict['ma18'][::-1][1:] # remove the lowest info
            stockInfoDict['ma50'] = stockInfoDict['ma50'][::-1][1:] # remove the lowest info
            stockInfoDict['k'] = stockInfoDict['k'][::-1][1:] # remove the lowest info
            stockInfoDict['d'] = stockInfoDict['d'][::-1][1:] # remove the lowest info
            stockInfoDict['amount'] = stockInfoDict['amount'][::-1]
            stockInfoDict['dailyAmount'] = stockInfoDict['dailyAmount'][::-1]

            # print(stockInfoDict['index'])

            # Add stockPrediction
            stock_prediction = False
            stock_prediction_add = 0
            if stock_prediction:
                cursor.execute('SELECT stockCode, stockPrediction FROM stockdata WHERE stockCode=%s ORDER BY stockDate DESC LIMIT 1', (stock))
                results = cursor.fetchall()
                for row in results:
                    stockInfoDict['index'].append(float(row[1]))

                xtickLabels.append('Next')
                stock_prediction_add = 1

            plt.setp(subplot1, xticks=[i for i in range(1, days+1+stock_prediction_add)], xticklabels=xtickLabels, xlim=[0, days+1+stock_prediction_add])#, ylabel='score')  # the last +1 is for stockPrediction
            # plt.xticks(rotation=10)
            for tick in subplot1.get_xticklabels():
                tick.set_rotation(20)
            pL11 = subplot1.plot([i for i in range(1, days+1+stock_prediction_add)], stockInfoDict['index'], '', label='stockIndex', zorder=10)  # the last +1 is for stockPrediction
            for xCor, yCor in zip([i for i in range(1, days+1+stock_prediction_add)], stockInfoDict['index']):  # the last +1 is for stockPrediction
                subplot1.text(xCor, yCor, str(yCor), weight='bold')
                # subplot1.text(xCor-0.2, yCor-0.06, str(yCor), weight='bold')
            pL12 = subplot1.plot(x, stockInfoDict['ma18'], '', label='stockMA18', zorder=10)
            # for xCor, yCor in zip(x, y2):
            #     subplot1.text(xCor, yCor+0.05, str(yCor), weight='bold')
            pL13 = subplot1.plot(x, stockInfoDict['ma50'], '', label='stockMA50', zorder=10)
            # for xCor, yCor in zip(x, y3):
            #     subplot1.text(xCor, yCor-0.05, str(yCor), weight='bold')
            subplot1.plot(0.1, lowestIndex, 'ro')
            subplot1.text(0.1, lowestIndex, str(lowestIndex), weight='bold')

            amountColor = []
            for amIndex in range(len(stockInfoDict['amount'])):
                if int(stockInfoDict['amount'][amIndex])<0:
                    amountColor.append('green')
                    stockInfoDict['amount'][amIndex] = -float(stockInfoDict['amount'][amIndex])
                else:
                    amountColor.append('red')

            dailyAmountColor = []
            for daiAmIndex in range(len(stockInfoDict['dailyAmount'])):
                if int(stockInfoDict['dailyAmount'][daiAmIndex]) < 0:
                    dailyAmountColor.append('green')
                    stockInfoDict['dailyAmount'][daiAmIndex] = -float(stockInfoDict['dailyAmount'][daiAmIndex])
                else:
                    dailyAmountColor.append('red')

            print(stockInfoDict['amount'])
            print(stockInfoDict['dailyAmount'])

            width = 0.5
            pLBar = subplot1.twinx()
            if max(stockInfoDict['dailyAmount'])*1.1!=0:
                maxAmount = max(stockInfoDict['dailyAmount'])*1.1
            else:
                maxAmount = 10000
            plt.setp(pLBar, xticks=[i for i in range(1, days+1+stock_prediction_add)], xticklabels=xtickLabels, xlim=[0, days+1+stock_prediction_add], ylim=[0, maxAmount])  # the last +1 is for stockPrediction

            print(stockInfoDict)

            # pLBar.bar(x, stockInfoDict['amount'], width, alpha=0.2, label='amount', color=amountColor, zorder=1)
            pLBar.bar(x - width / 2 + 0.1/2, stockInfoDict['amount'], width-0.1, alpha = 0.2, label='amount', color=amountColor)#, zorder=1)
            pLBar.bar(x + width / 2 - 0.1/2, stockInfoDict['dailyAmount'], width-0.1, alpha=0.2, label='dailyAmount', color=dailyAmountColor)#, zorder=1)
            for xCor, yCor in zip(x-width, stockInfoDict['amount']):
                pLBar.text(xCor, yCor, str(int(yCor)))
            for xCor, yCor in zip(x, stockInfoDict['dailyAmount']):
                pLBar.text(xCor, yCor, str(int(yCor)))

            # 'best'         : 0, (only implemented for axis legends)
            # 'upper right'  : 1,
            # 'upper left'   : 2,
            # 'lower left'   : 3,
            # 'lower right'  : 4,
            # 'right'        : 5,
            # 'center left'  : 6,
            # 'center right' : 7,
            # 'lower center' : 8,
            # 'upper center' : 9,
            # 'center'       : 10,

            h1, l1 = subplot1.get_legend_handles_labels()
            lgd = subplot1.legend(h1, l1, loc=8, fancybox=False, shadow=False, ncol=1) # http://matplotlib.org/users/legend_guide.html

            # subplot2
            plt.setp(subplot2,xticks=x, xticklabels='')#, ylabel='score')
            pL21 = subplot2.plot(x, stockInfoDict['k'], '', label='stockK', zorder=10)
            # for xCor, yCor in zip(x, y21):
            #     subplot2.text(xCor-0.08, yCor+1, str("%.2f" % (yCor)), weight='bold')
            pL22 = subplot2.plot(x, stockInfoDict['d'], '', label='stockD', zorder=10)
            # for xCor, yCor in zip(x, y22):
            #     subplot2.text(xCor-0.08, yCor-10, str("%.2f" % (yCor)), weight='bold')
            subplot2.axhline(y=20, color='g', linestyle='--')
            subplot2.axhline(y=80, color='r', linestyle='-.')

            h2, l2 = subplot2.get_legend_handles_labels()
            lgd2 = subplot2.legend(h2, l2, loc=3, fancybox=False, shadow=False, ncol=1) # http://matplotlib.org/users/legend_guide.html

            fig.tight_layout()
            fig.savefig('flask-stock-decken/static/stockDrawing_' + stock + '_' + stockCodeNames[stock][0] + '.png', bbox_inches='tight')
            plt.close(fig)
        #___ Draw diagrams

            # Save today's potential stocks in table 'stocktable'
            while True:
                try:
                    sql = "UPDATE stocktable SET stockIsTodayPotential=1 WHERE stockCode=%s"
                    cursor.execute(sql, (stock))
                    conn.commit()
                    break
                except:
                    print("Unexpected error:", sys.exc_info())

        conn.commit()
        cursor.close()
        conn.close()

        self.generate_potential_stocks_txt()

    def generate_potential_stocks_txt(self):
        try:
            # Execute the SQL command
            conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
            cursor = conn.cursor()
            cursor.execute('SELECT stockcode,stockName,stockIsYesterdayPotential,stockIsTodayPotential FROM stocktable WHERE stockIsYesterdayPotential=1 OR stockIsTodayPotential=1')
            # Fetch all the rows in a list of lists.
            stock_potential_info = {}
            if_potential_dict = {0: '不是', 1: '是'}
            results = cursor.fetchall()
            certificate_deposit_stocks = ['0050', '2330', '2412', '2105', '2596', '2823', '2885', '2891', '6218']
            for row in results:
                stock_potential_info[row[0]] = [row[1], if_potential_dict[int(row[2])], if_potential_dict[int(row[3])], '定存股' if row[0] in certificate_deposit_stocks else '']

        except:
            print("Error: unable to fecth data")

        cursor.close()
        conn.commit()
        conn.close()

        import json
        with open('flask-stock-decken/potential_stocks.json', 'w', encoding='utf8') as outfile:
            json.dump(stock_potential_info, outfile, ensure_ascii=False)

        # return stock_potential_info

    def getPttStockNewsComments(self,pages):
        pages = int(pages)
        ua = UserAgent()
        header = {'User-Agent': str(ua.random)}
        urlPttStock = 'https://www.ptt.cc/bbs/Stock/index.html'
        urlList = []
        successfulPages = 0

        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cur = conn.cursor()
        titleRemoved = []
        try:
            cur.execute('SELECT phraseRemoved FROM stockphraseremoved WHERE type="title"')
            results = cur.fetchall()
            for row in results:
                titleRemoved.append(row[0])
        except:
            print("Unexpected error:", sys.exc_info())

        phraseRemoved = []
        try:
            cur.execute('SELECT phraseRemoved FROM stockphraseremoved WHERE type<>"title"')
            results = cur.fetchall()
            for row in results:
                phraseRemoved.append(row[0])
        except:
            print("Unexpected error:", sys.exc_info())

        while successfulPages < pages:
            try:
                nowProxy = self.getProxy(0)
                proxies = {"http": "http://" + nowProxy}
                res = requests.get(urlPttStock, headers=header, proxies=proxies, timeout=10)#, proxies=proxies, timeout=self.timeout)
                if '批踢踢實業坊' not in res.text:
                    continue

                html = etree.HTML(res.text)
                result = html.xpath('//a/@href')
                previousPageUrl = ''
                for oneUrl in result:
                    splitUrl = oneUrl.split('/')
                    letters = collections.Counter(splitUrl[-1])
                    oneUrl = 'https://www.ptt.cc' + oneUrl.strip()
                    if letters['.']>=3 and oneUrl not in urlList: # filter out urls with less than three dots
                        urlList.append(oneUrl)
                    elif re.search('index[0-9]{4,}', splitUrl[-1]) and previousPageUrl=='':
                        previousPageUrl = oneUrl

                successfulPages += 1
                urlPttStock = previousPageUrl
                time.sleep(2)
            except:
                print("Unexpected error:", sys.exc_info())

        print(urlList)

        # urlList = ['https://www.ptt.cc/bbs/Stock/M.1519884995.A.23D.html']

        sql = "INSERT IGNORE INTO stocknewscomments (stockNCUrl,stockNCAuthor,stockNCTitle,stockNCContent,stockNCPosttime) VALUES (%s,%s,%s,%s,%s)"
        insertNewsCommentsArray = []

        # urlList = ['https://www.ptt.cc/bbs/Stock/M.1524483102.A.34E.html']
        for oneUrl in urlList:
            print(oneUrl)
            while True:
                ua = UserAgent()
                header = {'User-Agent': str(ua.random)}

                try: # for catching requests error
                    res = requests.get(oneUrl, headers=header, timeout=10)#, proxies=proxies, timeout=self.timeout)
                    if '批踢踢實業坊' not in res.text:
                        continue

                    html = etree.HTML(res.text)

                    # matchedPushUserIds = html.xpath('//span[contains(@class,"push-userid")]')
                    # pushUserIds = []
                    # for oneUser in matchedPushUserIds:
                    #     pushUserIds.append(oneUser.text.strip())
                    # print(pushUserIds)

                    try: #for catching invalid urls
                        result = html.xpath('//span[contains(@class,"article-meta-value")]')
                        #[0]=>author name, [1]=>Forum name(Stock), [2]=>title, [3]=>post time
                        authorName = result[0].text.strip().split(' ')[0]

                        title = result[2].text.strip()
                        removed = False
                        for titleRe in titleRemoved:
                            if titleRe in title:
                                removed = True
                                break
                        if removed:
                            break

                        monthNumber = dict((v,str(k).zfill(2)) for k,v in enumerate(calendar.month_abbr)) # {'': '00', 'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
                        timeStampSeparate = result[3].text.replace('  ',' ').split(' ') #because the first day of one month takes two spaces (Thu Mar  1 14:16:32 2018). ['Sun', 'Feb', '25', '16:37:51', '2018']
                        # print(result[3].text)
                        # print(timeStampSeparate)
                        if int(timeStampSeparate[4]) > datetime.datetime.now().year:
                            timeStampSeparate[4] = str(datetime.datetime.now().year)
                        timeStamp = timeStampSeparate[4] + '-' + monthNumber[timeStampSeparate[1]] + '-' + timeStampSeparate[2].zfill(2) + ' ' + timeStampSeparate[3] #zfill(2) to transform 1 to 01
                    except Exception as e:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print(exc_type, fname, exc_tb.tb_lineno)
                        break

                    reGetContent = re.compile(r'<span class="article-meta-value">.*?[\d]{4}<\/span><\/div>(.*?)--.*?<span class="', re.S|re.UNICODE)
                    content = ''
                    for content in reGetContent.findall(res.text):
                        content = content.strip()

                    #main post
                    for phrase in phraseRemoved:
                        content = content.replace(phrase,'')
                    content = re.sub(r"\<a.*?\<\/a>", "", content) # remove link texts. re.sub(r"\<a.*?href.*?>|<\/a>", "", content)

                    insertNewsCommentsArray.append((oneUrl,authorName,title,content,timeStamp))
                    while True:
                            try:
                                print((oneUrl,authorName,title,content,timeStamp))
                                cur.execute(sql,((oneUrl+'?main',authorName,title,content,timeStamp)))
                                # cur.close()
                                conn.commit()
                                # conn.close()
                                break
                            except:
                                print("Unexpected error:", sys.exc_info())

                    #pushed comments
                    # reGetComments = re.compile(r'">:(.*?)[\s]*?<.*?([\d]{2}\/[\d]{2}\s+[\d]{2}:[\d]{2})', re.S|re.UNICODE)
                    # reGetComments = re.compile(r'[\/\w\s"=-]+>([\w_=-]+).*?<[\/\w\s"=-]+><[\w\s"=-]+">:(.*?)[\s]*?<\/span.*?([\d]{2}\/[\d]{2}\s+[\d]{2}:[\d]{2})', re.S|re.UNICODE)
                    reGetComments = re.compile(r'userid.*?>([\w_=-]+).*?<[\/\w\s"=-]+><[\w\s"=-]+">:(.*?)[\s]*?<\/span.*?([\d]{2}\/[\d]{2}\s+[\d]{2}:[\d]{2})', re.S|re.UNICODE)
                    pushedUserIdCounter = 0
                    # print(res.text)
                    for userId,comment,commentTimeStamp in reGetComments.findall(res.text):
                        comment = re.sub(r'<a.*?>', '', comment.strip()).replace('</a>','')
                        commentTimeStampTemp = commentTimeStamp
                        commentTimeStamp = timeStamp.split('-')[0] + '-' + commentTimeStamp.replace('/','-').strip()
                        # print(timeStamp)
                        # print(commentTimeStamp)
                        # print(commentTimeStamp+':59')
                        if commentTimeStamp+':59' < timeStamp: # this means the time of the comment is the next year (:59 is to avoid accidentally jumping to the next year)
                            if int(timeStamp.split('-')[0])+1 > datetime.datetime.now().year:
                                commentTimeStamp = str(datetime.datetime.now().year) + '-' + commentTimeStampTemp.replace('/','-').strip()
                            else:
                                commentTimeStamp = str(int(timeStamp.split('-')[0])+1) + '-' + commentTimeStampTemp.replace('/','-').strip()
                        # print(str(comment) + '_' + str(commentTimeStamp))
                        insertNewsCommentsArray.append((oneUrl,userId,title,comment,commentTimeStamp + ':' + str(pushedUserIdCounter%60).zfill(2))) # add sequential seconds to prevent duplicates

                        while True:
                            try:
                                print((oneUrl,userId,title,comment,commentTimeStamp + ':' + str(pushedUserIdCounter%60).zfill(2)))
                                for phrase in phraseRemoved:
                                    comment = comment.replace(phrase, '')
                                comment = re.sub(r"\<a.*?\<\/a>", "", comment)  # remove link texts
                                cur.execute(sql,(oneUrl,userId,title,comment,commentTimeStamp + ':' + str(pushedUserIdCounter%60).zfill(2)))
                                # cur.close()
                                conn.commit()
                                # conn.close()
                                break
                            except:
                                print("Unexpected error:", sys.exc_info())

                        pushedUserIdCounter += 1
                    break
                except:
                    print("Unexpected error:", sys.exc_info())

            time.sleep(2)

        # print(insertNewsCommentsArray)
        # if insertNewsCommentsArray:
        #     while True:
        #         try:
        #             cur.executemany(sql,insertNewsCommentsArray)
        #             cur.close()
        #             conn.commit()
        #             conn.close()
        #             break
        #         except:
        #             print("Unexpected error:", sys.exc_info())

    def getStockNameToCode(self):
        stockNameToCode = {}
        try:
            # Execute the SQL command
            conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
            cursor = conn.cursor()
            cursor.execute("SELECT stockCode,TRIM(TRAILING ';' FROM CONCAT_WS(';',stockName,stockOthername)) AS stockAllnames,stockName FROM stocktable")
            # Fetch all the rows in a list of lists.
            results = cursor.fetchall()
            for row in results:
                stockNameToCode[row[0]] = row[2]
                for stockName in row[1].split(';'):
                    stockNameToCode[stockName] = row[2]

            # print(stockNameToCode)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        return stockNameToCode

    def getVocabulary(self,tableName):
        vocabularyList = list()
        try:
            # Execute the SQL command
            conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
            cursor = conn.cursor()
            cursor.execute('SELECT GROUP_CONCAT(CONCAT(`word_trad`) SEPARATOR ";") AS positivePhrase FROM ' + tableName)
            # Fetch all the rows in a list of lists.
            results = cursor.fetchall()
            for row in results:
                for vocabulary in row[0].split(';'):
                    vocabularyList.append(vocabulary)

            # print(vocabularyList)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        return vocabularyList

    def FullToHalf(self,s):
        n = []
        # s = s.decode('utf-8')
        for char in s:
            num = ord(char)
            if num == 0x3000:
                num = 32
            elif 0xFF01 <= num <= 0xFF5E:
                num -= 0xfee0
            num = chr(num)
            n.append(num)
        return ''.join(n)

    def computeStockSentiment(self,content,stockUrl):
        jieba.set_dictionary('jieba/dict.txt.big')
        # content = open('jieba/newsComment.txt', 'r', encoding="utf-8").read()
        # stockUrl = 'https://www.ptt.cc/bbs/Stock/M.1520011815.A.6B1.html'
        mainSentiment = None

        # Also get the main news to append it to the current one if not finished computing sentiments yet
        if 'main' not in stockUrl:
            try:
                # Execute the SQL command
                conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
                cursor = conn.cursor()
                cursor.execute("SELECT stockNCContent,stockNCSentiment FROM stocknewscomments WHERE stockNCUrl LIKE '%" + stockUrl + "?main%'")
                # Fetch all the rows in a list of lists.
                results = cursor.fetchall()
                for row in results:
                    if row[1]=='':
                        content += row[0]
                    else:
                        mainSentiment = json.loads(row[1])
            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)

        positiveVocabularyList = self.getVocabulary('stockpositivevocabulary')
        negativeVocabularyList = self.getVocabulary('stocknegativevocabulary')

        stockNameToCode = self.getStockNameToCode()
        # Add phrases to jieba
        for stockName in stockNameToCode:
            jieba.suggest_freq(stockName, True)
        words = jieba.cut(self.FullToHalf(content), cut_all=False)

        currentStock = '' # remember which stock the comment points to
        accumulatedNonStockSentiment = 0 # if no stock is identified yet, sum sentiments here
        stockcodeToSentiment = dict()
        stockNameCounter = dict()
        for word in words:
            wordIsStock = False
            for stockName in stockNameToCode:
                if stockName in word:
                    wordIsStock = True
                    currentStock = stockNameToCode[stockName]
                    stockNameCounter[currentStock] = stockNameCounter.get(currentStock, 0) + 1

                    if currentStock not in stockcodeToSentiment:
                        stockcodeToSentiment[currentStock] = 0

                    # print(stockName, stockNameToCode[stockName], word)
                    break

            if not wordIsStock and currentStock=='':
                if word in positiveVocabularyList:
                    accumulatedNonStockSentiment += 1
                elif word in negativeVocabularyList:
                    accumulatedNonStockSentiment -= 1

                # print('compute the sentiment and add it to accumulatedNonStockSentiment')
            elif not wordIsStock and currentStock!='':
                if word in positiveVocabularyList:
                    stockcodeToSentiment[currentStock] += 1
                elif word in negativeVocabularyList:
                    stockcodeToSentiment[currentStock] -= 1

                # print('compute the sentiment and add it to the stock\'s sentiment')

        # add the unknown sentiment to the most frequently-mentioned stock
        stockCounterSorted = sorted(stockNameCounter.items(), key=itemgetter(1), reverse=True)
        for oneItem in stockcodeToSentiment:
            if oneItem==stockCounterSorted[0][0]:
                stockcodeToSentiment[oneItem] = int(stockcodeToSentiment[oneItem])+int(accumulatedNonStockSentiment)
                break

        # add the sentiment of the main news to the current one
        if mainSentiment:
            for stock in mainSentiment:
                stockcodeToSentiment[stock] = stockcodeToSentiment.get(stock, mainSentiment[stock]) + mainSentiment[stock]

        sentimentSorted = sorted(stockcodeToSentiment.items(), key=itemgetter(1), reverse=True)
        toJson = '{'
        for oneItem in sentimentSorted:
            # if highest:
            #     toJson += '\'' + str(oneItem[0]) + '\'' + ':' + str(int(oneItem[1])+int(accumulatedNonStockSentiment)) + ','
            #     highest = False
            # else:
            toJson += '\"' + str(oneItem[0]) + '\"' + ':' + str(oneItem[1]) + ','
        toJson = toJson.rstrip(',')
        toJson += '}'

        print(toJson)
        return toJson

    def getChineseCharacter(self,input):
        chCharacters = ''
        for n in re.findall('[\u4e00-\u9fff]+', input):
            chCharacters += n

        return chCharacters

    def listToNGram(self,inputList, ngram):
        return [inputList[i:i + ngram] for i in range(0, len(inputList) - ngram + 1)]

    def listToFreqdict(self,inputList):
        outputDict = dict()
        for i in range(len(inputList)):
            # outputDict[tuple(inputList[i])] = outputDict.get(tuple(inputList[i]), 0) + 1
            outputDict[''.join(inputList[i])] = outputDict.get(''.join(inputList[i]), 0) + 1
        chfreqsorted = sorted(outputDict.items(), key=itemgetter(1), reverse=True)
        toJson = '{'
        for oneItem in chfreqsorted:
            toJson += '\"' + str(oneItem[0]) + '\"' + ':' + str(oneItem[1]) + ','
        toJson = toJson.rstrip(',')
        toJson += '}'

        # return chfreqsorted
        return toJson

    def computeNGramfreq(self,dummy):
        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cursor = conn.cursor()
        ngramArray = []

        try:
            # Execute the SQL command
            # cursor.execute("SELECT sid,CONCAT_WS('',stockNCTitle,stockNCContent) AS titleContent,stockNCUrl FROM stocknewscomments WHERE stockNC2gramfreq=''")
            cursor.execute("SELECT sid,stockNCContent,stockNCUrl FROM stocknewscomments WHERE stockNC2gramfreq=''")
            # Fetch all the rows in a list of lists.
            results = cursor.fetchall()
            counter = 0
            for row in results:
                counter += 1
                chCharacters = self.getChineseCharacter(row[1])
                chlist = [ch for ch in chCharacters]

                stockSentiment = self.computeStockSentiment(row[1],row[2])
                twoGramDict = self.listToFreqdict( self.listToNGram(chlist,2) )
                threeGramDict = self.listToFreqdict( self.listToNGram(chlist, 3) )
                fourGramDict = self.listToFreqdict( self.listToNGram(chlist, 4) )

                # print(stockSentiment,twoGramDict,threeGramDict,fourGramDict)
                ngramArray.append((str(stockSentiment),str(twoGramDict),str(threeGramDict),str(fourGramDict),row[0]))

                while counter >= 100:
                    try:
                        sql = "UPDATE stocknewscomments SET stockNCSentiment=%s,stockNC2gramfreq=%s,stockNC3gramfreq=%s,stockNC4gramfreq=%s WHERE sid=%s"
                        # print(ngramArray)
                        cursor.executemany(sql, ngramArray)
                        conn.commit()
                        counter = 0
                        ngramArray = []
                        break
                    except:
                        print("Unexpected error:", sys.exc_info())
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        # update remaining data
        while True:
            try:
                sql = "UPDATE stocknewscomments SET stockNCSentiment=%s,stockNC2gramfreq=%s,stockNC3gramfreq=%s,stockNC4gramfreq=%s WHERE sid=%s"
                # print(ngramArray)
                cursor.executemany(sql, ngramArray)
                conn.commit()
                counter = 0
                ngramArray = []
                break
            except:
                print("Unexpected error:", sys.exc_info())

        cursor.close()
        conn.close()

    def aggregateGramfreq(self,dummy):
        conn = pymysql.connect(host='192.168.2.55', port=3306, user='root', passwd='89787198', db='stockevaluation', charset="utf8")
        cursor = conn.cursor()

        # get the previous stockHandledCommentid
        try:
            # Execute the SQL command
            cursor.execute("SELECT stockHandledCommentid,stock2Gram,stock3Gram,stock4Gram FROM stockngramfreq ORDER BY stockHandledCommentid DESC LIMIT 1")
            results = cursor.fetchall()
            stockHandledCommentid = 0
            gramsDict = {2: {}, 3: {}, 4: {}}
            for row in results:
                if int(row[0]) > 0:
                    stockHandledCommentid = int(row[0])

                    for gramNum in range(2, 5):
                        gramJson = json.loads(row[gramNum - 1])
                        for key in gramJson:
                            gramsDict[gramNum][key] = gramJson[key]
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        try:
            # Execute the SQL command
            cursor.execute("SELECT sid,stockNC2gramfreq,stockNC3gramfreq,stockNC4gramfreq FROM stocknewscomments WHERE sid > %s AND stockNC2gramfreq <> '' AND stockNC2gramfreq <> '{}'",(stockHandledCommentid))
            # Fetch all the rows in a list of lists.
            results = cursor.fetchall()
            # gramsDict = {2:{},3:{},4:{}}
            maxSid = -1
            for row in results:
                if int(row[0]) > maxSid:
                    maxSid = int(row[0])

                for gramNum in range(2,5):
                    gramJson = json.loads(row[gramNum-1])

                    for key in gramJson:
                        gramsDict[gramNum][key] = gramsDict[gramNum].get(key,0) + int(gramJson[key])

            # sort, convert to json and insert into stockngramfreq
            gramsJsonDict = {2: '', 3: '', 4: ''}
            for gramNum in range(2,5):
                gramSorted = sorted(gramsDict[gramNum].items(), key=itemgetter(1), reverse=True)
                gramToJson = '{'
                for oneItem in gramSorted:
                    # if highest:
                    #     toJson += '\'' + str(oneItem[0]) + '\'' + ':' + str(int(oneItem[1])+int(accumulatedNonStockSentiment)) + ','
                    #     highest = False
                    # else:
                    gramToJson += '\"' + str(oneItem[0]) + '\"' + ':' + str(oneItem[1]) + ','
                gramToJson = gramToJson.rstrip(',')
                gramToJson += '}'

                gramsJsonDict[gramNum] = gramToJson
                print(gramsJsonDict[gramNum])

                # print(sorted(gramsDict[2].items(), key=itemgetter(1), reverse=True))
                # print(sorted(gramsDict[3].items(), key=itemgetter(1), reverse=True))
                # print(sorted(gramsDict[4].items(), key=itemgetter(1), reverse=True))

            cursor.execute('TRUNCATE TABLE stockngramfreq')
            print(maxSid)
            insert = "INSERT IGNORE INTO stockngramfreq (stock2Gram, stock3Gram, stock4Gram, stockHandledCommentid) VALUES (%s, %s, %s, %s)"
            cursor.execute(insert, (gramsJsonDict[2], gramsJsonDict[3], gramsJsonDict[4], maxSid))
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        cursor.close()
        conn.commit()
        conn.close()

    @staticmethod
    def make_car_sound():
        print('VRooooommmm!')

    @abstractmethod
    def vehicle_type(self):
        """"Return a string representing the type of vehicle this is."""
        pass